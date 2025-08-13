"""
Advanced Reporting Service
Generates comprehensive reports in multiple formats (PDF, CSV, JSON, Excel)
with data visualization, executive summaries, and actionable insights.
"""

import asyncio
import json
import csv
import io
import base64
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional, Union, BinaryIO
from dataclasses import dataclass, asdict
from enum import Enum
import tempfile
import os

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie

# Excel generation
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from .trend_analysis import trend_analysis_engine
from .social_media_service import social_media_service
from .ai_content_analyzer import ai_content_analyzer

class ReportFormat(Enum):
    PDF = "pdf"
    CSV = "csv"
    JSON = "json"
    EXCEL = "xlsx"
    HTML = "html"

class ReportType(Enum):
    EXECUTIVE_SUMMARY = "executive_summary"
    DETAILED_ANALYSIS = "detailed_analysis"
    TREND_REPORT = "trend_report"
    COMPETITIVE_LANDSCAPE = "competitive_landscape"
    CONTENT_ANALYSIS = "content_analysis"
    SOCIAL_MEDIA_REPORT = "social_media_report"
    CUSTOM = "custom"

@dataclass
class ReportConfig:
    """Configuration for report generation."""
    report_type: ReportType
    format: ReportFormat
    title: str
    description: str
    include_charts: bool = True
    include_raw_data: bool = False
    date_range_days: int = 30
    competitor_ids: List[str] = None
    industry: str = None
    custom_sections: List[str] = None

@dataclass
class ReportSection:
    """Individual report section."""
    title: str
    content: str
    data: Dict[str, Any] = None
    chart_data: Dict[str, Any] = None
    section_type: str = "text"  # text, table, chart, image

@dataclass
class GeneratedReport:
    """Generated report result."""
    report_id: str
    config: ReportConfig
    sections: List[ReportSection]
    file_path: str
    file_size_bytes: int
    generation_time_ms: int
    created_at: datetime
    metadata: Dict[str, Any]

class AdvancedReportingService:
    """Advanced reporting service with multiple export formats."""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._create_custom_styles()
        
    def _create_custom_styles(self):
        """Create custom styles for PDF generation."""
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.Color(0.2, 0.3, 0.6)
        ))
        
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=16,
            spaceBefore=20,
            spaceAfter=12,
            textColor=colors.Color(0.3, 0.4, 0.7)
        ))
        
        self.styles.add(ParagraphStyle(
            name='Insight',
            parent=self.styles['Normal'],
            fontSize=11,
            leftIndent=20,
            rightIndent=20,
            spaceBefore=6,
            spaceAfter=6,
            backColor=colors.Color(0.95, 0.97, 1.0)
        ))

    async def generate_report(self, config: ReportConfig) -> GeneratedReport:
        """Generate a comprehensive report based on configuration."""
        start_time = datetime.utcnow()
        
        # Collect data based on report type
        report_data = await self._collect_report_data(config)
        
        # Generate report sections
        sections = await self._generate_report_sections(config, report_data)
        
        # Generate the actual report file
        file_path = await self._generate_report_file(config, sections)
        
        # Get file size
        file_size = os.path.getsize(file_path)
        
        end_time = datetime.utcnow()
        generation_time = int((end_time - start_time).total_seconds() * 1000)
        
        return GeneratedReport(
            report_id=self._generate_report_id(),
            config=config,
            sections=sections,
            file_path=file_path,
            file_size_bytes=file_size,
            generation_time_ms=generation_time,
            created_at=end_time,
            metadata={
                "total_sections": len(sections),
                "data_points": sum(len(s.data) if s.data else 0 for s in sections),
                "format": config.format.value
            }
        )

    async def _collect_report_data(self, config: ReportConfig) -> Dict[str, Any]:
        """Collect all necessary data for the report."""
        data = {}
        
        # Base date range
        end_date = datetime.utcnow()
        start_date = end_date - timedelta(days=config.date_range_days)
        
        try:
            # Collect trends data
            content_trends = await trend_analysis_engine.analyze_content_trends(
                industry=config.industry,
                days_back=config.date_range_days
            )
            
            social_trends = await trend_analysis_engine.analyze_social_media_trends(
                industry=config.industry,
                days_back=config.date_range_days
            )
            
            # Generate market insights
            all_trends = content_trends + social_trends
            insights = await trend_analysis_engine.generate_market_insights(
                trends=all_trends,
                industry=config.industry
            )
            
            data.update({
                "content_trends": content_trends,
                "social_trends": social_trends,
                "market_insights": insights,
                "date_range": {"start": start_date, "end": end_date}
            })
            
            # Add competitor-specific data if specified
            if config.competitor_ids:
                competitor_data = {}
                for comp_id in config.competitor_ids:
                    try:
                        # Get social analytics
                        social_analytics = await social_media_service.get_social_analytics(comp_id)
                        competitor_data[comp_id] = {
                            "social_analytics": social_analytics
                        }
                    except Exception as e:
                        competitor_data[comp_id] = {"error": str(e)}
                
                data["competitors"] = competitor_data
            
        except Exception as e:
            data["error"] = f"Data collection error: {str(e)}"
        
        return data

    async def _generate_report_sections(self, config: ReportConfig, data: Dict[str, Any]) -> List[ReportSection]:
        """Generate report sections based on type and data."""
        sections = []
        
        # Executive Summary (always included)
        sections.append(await self._generate_executive_summary(config, data))
        
        if config.report_type == ReportType.EXECUTIVE_SUMMARY:
            sections.extend([
                await self._generate_key_metrics_section(data),
                await self._generate_top_insights_section(data),
                await self._generate_recommendations_section(data)
            ])
        
        elif config.report_type == ReportType.DETAILED_ANALYSIS:
            sections.extend([
                await self._generate_trends_analysis_section(data),
                await self._generate_competitor_analysis_section(data),
                await self._generate_content_performance_section(data),
                await self._generate_social_media_section(data),
                await self._generate_strategic_insights_section(data)
            ])
        
        elif config.report_type == ReportType.TREND_REPORT:
            sections.extend([
                await self._generate_trending_topics_section(data),
                await self._generate_emerging_patterns_section(data),
                await self._generate_trend_predictions_section(data)
            ])
        
        elif config.report_type == ReportType.COMPETITIVE_LANDSCAPE:
            sections.extend([
                await self._generate_competitor_overview_section(data),
                await self._generate_market_positioning_section(data),
                await self._generate_competitive_gaps_section(data)
            ])
        
        elif config.report_type == ReportType.CONTENT_ANALYSIS:
            sections.extend([
                await self._generate_content_quality_section(data),
                await self._generate_topic_analysis_section(data),
                await self._generate_content_gaps_section(data)
            ])
        
        elif config.report_type == ReportType.SOCIAL_MEDIA_REPORT:
            sections.extend([
                await self._generate_social_overview_section(data),
                await self._generate_engagement_analysis_section(data),
                await self._generate_viral_content_section(data)
            ])
        
        # Add custom sections if specified
        if config.custom_sections:
            for section_name in config.custom_sections:
                sections.append(await self._generate_custom_section(section_name, data))
        
        return sections

    async def _generate_report_file(self, config: ReportConfig, sections: List[ReportSection]) -> str:
        """Generate the actual report file in the specified format."""
        if config.format == ReportFormat.PDF:
            return await self._generate_pdf_report(config, sections)
        elif config.format == ReportFormat.CSV:
            return await self._generate_csv_report(config, sections)
        elif config.format == ReportFormat.JSON:
            return await self._generate_json_report(config, sections)
        elif config.format == ReportFormat.EXCEL:
            return await self._generate_excel_report(config, sections)
        elif config.format == ReportFormat.HTML:
            return await self._generate_html_report(config, sections)
        else:
            raise ValueError(f"Unsupported report format: {config.format}")

    async def _generate_pdf_report(self, config: ReportConfig, sections: List[ReportSection]) -> str:
        """Generate PDF report."""
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        temp_file.close()
        
        # Create PDF document
        doc = SimpleDocTemplate(temp_file.name, pagesize=A4)
        story = []
        
        # Title page
        story.append(Paragraph(config.title, self.styles['CustomTitle']))
        story.append(Paragraph(config.description, self.styles['Normal']))
        story.append(Spacer(1, 20))
        story.append(Paragraph(f"Generated on: {datetime.utcnow().strftime('%B %d, %Y')}", self.styles['Normal']))
        story.append(Spacer(1, 40))
        
        # Add sections
        for section in sections:
            story.append(Paragraph(section.title, self.styles['SectionHeader']))
            story.append(Paragraph(section.content, self.styles['Normal']))
            
            # Add data tables if present
            if section.data:
                table_data = self._format_data_for_table(section.data)
                if table_data:
                    table = Table(table_data)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 14),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    story.append(table)
            
            story.append(Spacer(1, 20))
        
        # Build PDF
        doc.build(story)
        return temp_file.name

    async def _generate_csv_report(self, config: ReportConfig, sections: List[ReportSection]) -> str:
        """Generate CSV report."""
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.csv', mode='w', newline='')
        
        writer = csv.writer(temp_file)
        
        # Header
        writer.writerow(['Report', config.title])
        writer.writerow(['Generated', datetime.utcnow().isoformat()])
        writer.writerow([])
        
        # Sections
        for section in sections:
            writer.writerow(['Section', section.title])
            writer.writerow(['Content', section.content])
            
            if section.data:
                # Convert data to rows
                for key, value in section.data.items():
                    if isinstance(value, (list, dict)):
                        writer.writerow([key, json.dumps(value)])
                    else:
                        writer.writerow([key, str(value)])
            
            writer.writerow([])
        
        temp_file.close()
        return temp_file.name

    async def _generate_json_report(self, config: ReportConfig, sections: List[ReportSection]) -> str:
        """Generate JSON report."""
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.json', mode='w')
        
        report_data = {
            "metadata": {
                "title": config.title,
                "description": config.description,
                "report_type": config.report_type.value,
                "generated_at": datetime.utcnow().isoformat(),
                "date_range_days": config.date_range_days
            },
            "sections": [asdict(section) for section in sections]
        }
        
        json.dump(report_data, temp_file, indent=2, default=str)
        temp_file.close()
        return temp_file.name

    async def _generate_excel_report(self, config: ReportConfig, sections: List[ReportSection]) -> str:
        """Generate Excel report with multiple sheets."""
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()
        
        workbook = Workbook()
        
        # Summary sheet
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        
        # Header styling
        header_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        summary_sheet['A1'] = config.title
        summary_sheet['A1'].font = Font(bold=True, size=18)
        summary_sheet['A3'] = config.description
        summary_sheet['A5'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Add sections to separate sheets
        row = 7
        for i, section in enumerate(sections):
            # Add section to summary
            summary_sheet[f'A{row}'] = section.title
            summary_sheet[f'A{row}'].font = header_font
            summary_sheet[f'B{row}'] = section.content[:100] + "..." if len(section.content) > 100 else section.content
            row += 1
            
            # Create detailed sheet for section if it has data
            if section.data:
                sheet_name = section.title[:30]  # Excel sheet name limit
                sheet = workbook.create_sheet(title=sheet_name)
                
                # Add section data
                sheet['A1'] = section.title
                sheet['A1'].font = header_font
                sheet['A3'] = section.content
                
                # Add data table
                data_row = 5
                for key, value in section.data.items():
                    sheet[f'A{data_row}'] = key
                    sheet[f'B{data_row}'] = str(value) if not isinstance(value, (list, dict)) else json.dumps(value)
                    data_row += 1
        
        workbook.save(temp_file.name)
        return temp_file.name

    async def _generate_html_report(self, config: ReportConfig, sections: List[ReportSection]) -> str:
        """Generate HTML report."""
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.html', mode='w')
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>{config.title}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; }}
                .header {{ background-color: #f0f2f5; padding: 20px; border-radius: 8px; }}
                .section {{ margin: 20px 0; padding: 15px; border-left: 4px solid #4472C4; }}
                .data-table {{ border-collapse: collapse; width: 100%; margin: 10px 0; }}
                .data-table th, .data-table td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                .data-table th {{ background-color: #4472C4; color: white; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{config.title}</h1>
                <p>{config.description}</p>
                <p>Generated on: {datetime.utcnow().strftime('%B %d, %Y at %H:%M:%S')}</p>
            </div>
        """
        
        for section in sections:
            html_content += f"""
            <div class="section">
                <h2>{section.title}</h2>
                <p>{section.content}</p>
            """
            
            if section.data:
                html_content += '<table class="data-table"><thead><tr><th>Key</th><th>Value</th></tr></thead><tbody>'
                for key, value in section.data.items():
                    value_str = str(value) if not isinstance(value, (list, dict)) else json.dumps(value)
                    html_content += f'<tr><td>{key}</td><td>{value_str}</td></tr>'
                html_content += '</tbody></table>'
            
            html_content += '</div>'
        
        html_content += '</body></html>'
        
        temp_file.write(html_content)
        temp_file.close()
        return temp_file.name

    # Section generators
    async def _generate_executive_summary(self, config: ReportConfig, data: Dict[str, Any]) -> ReportSection:
        """Generate executive summary section."""
        trends_count = len(data.get('content_trends', [])) + len(data.get('social_trends', []))
        insights_count = len(data.get('market_insights', []))
        
        content = f"""
        This {config.report_type.value.replace('_', ' ').title()} report provides comprehensive competitive intelligence 
        analysis for the period of {config.date_range_days} days ending {datetime.utcnow().strftime('%B %d, %Y')}.
        
        Key findings include {trends_count} identified trends and {insights_count} strategic insights that provide 
        actionable intelligence for competitive positioning and market opportunities.
        """
        
        return ReportSection(
            title="Executive Summary",
            content=content.strip(),
            data={
                "trends_identified": trends_count,
                "insights_generated": insights_count,
                "report_period_days": config.date_range_days,
                "generation_date": datetime.utcnow().isoformat()
            }
        )

    async def _generate_key_metrics_section(self, data: Dict[str, Any]) -> ReportSection:
        """Generate key metrics section."""
        content_trends = data.get('content_trends', [])
        social_trends = data.get('social_trends', [])
        insights = data.get('market_insights', [])
        
        # Calculate metrics
        strong_trends = [t for t in content_trends + social_trends if getattr(t, 'strength', '') in ['strong', 'viral']]
        high_impact_insights = [i for i in insights if getattr(i, 'impact_level', '') in ['high', 'critical']]
        
        content = f"""
        Key performance indicators and metrics from the analysis period:
        
        • Total Trends Identified: {len(content_trends + social_trends)}
        • Strong/Viral Trends: {len(strong_trends)}
        • Strategic Insights: {len(insights)}
        • High-Impact Insights: {len(high_impact_insights)}
        • Content Trends: {len(content_trends)}
        • Social Media Trends: {len(social_trends)}
        """
        
        return ReportSection(
            title="Key Metrics",
            content=content,
            data={
                "total_trends": len(content_trends + social_trends),
                "strong_trends": len(strong_trends),
                "total_insights": len(insights),
                "high_impact_insights": len(high_impact_insights),
                "content_trends": len(content_trends),
                "social_trends": len(social_trends)
            }
        )

    async def _generate_top_insights_section(self, data: Dict[str, Any]) -> ReportSection:
        """Generate top insights section."""
        insights = data.get('market_insights', [])
        
        # Sort by impact level and confidence
        top_insights = sorted(insights, 
                            key=lambda x: (getattr(x, 'impact_level', ''), getattr(x, 'confidence', 0)), 
                            reverse=True)[:5]
        
        content = "Top strategic insights from the analysis:\n\n"
        for i, insight in enumerate(top_insights, 1):
            content += f"{i}. {getattr(insight, 'title', 'Insight')}\n"
            content += f"   Impact: {getattr(insight, 'impact_level', 'Unknown')}\n"
            content += f"   {getattr(insight, 'description', 'No description available')}\n\n"
        
        return ReportSection(
            title="Top Strategic Insights",
            content=content,
            data={
                "insights": [
                    {
                        "title": getattr(insight, 'title', ''),
                        "impact_level": getattr(insight, 'impact_level', ''),
                        "confidence": getattr(insight, 'confidence', 0),
                        "description": getattr(insight, 'description', '')
                    }
                    for insight in top_insights
                ]
            }
        )

    async def _generate_recommendations_section(self, data: Dict[str, Any]) -> ReportSection:
        """Generate recommendations section."""
        insights = data.get('market_insights', [])
        
        recommendations = []
        for insight in insights:
            recs = getattr(insight, 'recommendations', [])
            recommendations.extend(recs)
        
        # Get unique recommendations
        unique_recs = list(set(recommendations))[:10]
        
        content = "Strategic recommendations based on competitive intelligence analysis:\n\n"
        for i, rec in enumerate(unique_recs, 1):
            content += f"{i}. {rec}\n"
        
        return ReportSection(
            title="Strategic Recommendations",
            content=content,
            data={"recommendations": unique_recs}
        )

    # Additional section generators would go here...
    async def _generate_trends_analysis_section(self, data: Dict[str, Any]) -> ReportSection:
        """Generate detailed trends analysis."""
        content_trends = data.get('content_trends', [])
        social_trends = data.get('social_trends', [])
        
        content = f"""
        Comprehensive trend analysis covering {len(content_trends)} content trends and {len(social_trends)} social media trends.
        
        Content trends show patterns in competitor content strategy, topic focus, and engagement approaches.
        Social media trends reveal platform-specific patterns, viral content characteristics, and audience engagement strategies.
        """
        
        trend_data = {}
        for trend in content_trends + social_trends:
            trend_data[getattr(trend, 'title', 'Unknown')] = {
                'type': getattr(trend, 'trend_type', 'unknown'),
                'strength': getattr(trend, 'strength', 'unknown'),
                'confidence': getattr(trend, 'confidence', 0)
            }
        
        return ReportSection(
            title="Trends Analysis",
            content=content,
            data=trend_data
        )

    async def _generate_competitor_analysis_section(self, data: Dict[str, Any]) -> ReportSection:
        """Generate competitor analysis section."""
        competitors = data.get('competitors', {})
        
        content = f"Analysis of {len(competitors)} competitors across multiple dimensions including content strategy, social media presence, and market positioning."
        
        return ReportSection(
            title="Competitor Analysis",
            content=content,
            data=competitors
        )

    async def _generate_content_performance_section(self, data: Dict[str, Any]) -> ReportSection:
        """Generate content performance section."""
        content = "Content performance analysis across competitors, including quality metrics, engagement patterns, and topic effectiveness."
        
        return ReportSection(
            title="Content Performance Analysis",
            content=content,
            data={"note": "Detailed content analysis would be populated from AI content analyzer"}
        )

    async def _generate_social_media_section(self, data: Dict[str, Any]) -> ReportSection:
        """Generate social media analysis section."""
        social_trends = data.get('social_trends', [])
        
        content = f"Social media landscape analysis covering {len(social_trends)} identified trends across platforms."
        
        return ReportSection(
            title="Social Media Analysis",
            content=content,
            data={"social_trends_count": len(social_trends)}
        )

    async def _generate_strategic_insights_section(self, data: Dict[str, Any]) -> ReportSection:
        """Generate strategic insights section."""
        insights = data.get('market_insights', [])
        
        content = f"Strategic market insights and competitive intelligence findings based on {len(insights)} analyzed patterns."
        
        return ReportSection(
            title="Strategic Market Insights",
            content=content,
            data={"insights_count": len(insights)}
        )

    # Additional placeholder sections
    async def _generate_trending_topics_section(self, data: Dict[str, Any]) -> ReportSection:
        return ReportSection("Trending Topics", "Analysis of trending topics and themes", {})
    
    async def _generate_emerging_patterns_section(self, data: Dict[str, Any]) -> ReportSection:
        return ReportSection("Emerging Patterns", "Identification of emerging market patterns", {})
    
    async def _generate_trend_predictions_section(self, data: Dict[str, Any]) -> ReportSection:
        return ReportSection("Trend Predictions", "Predictive analysis of future trends", {})
    
    async def _generate_competitor_overview_section(self, data: Dict[str, Any]) -> ReportSection:
        return ReportSection("Competitor Overview", "Comprehensive competitor landscape overview", {})
    
    async def _generate_market_positioning_section(self, data: Dict[str, Any]) -> ReportSection:
        return ReportSection("Market Positioning", "Analysis of competitive positioning strategies", {})
    
    async def _generate_competitive_gaps_section(self, data: Dict[str, Any]) -> ReportSection:
        return ReportSection("Competitive Gaps", "Identification of market gaps and opportunities", {})
    
    async def _generate_content_quality_section(self, data: Dict[str, Any]) -> ReportSection:
        return ReportSection("Content Quality Analysis", "AI-powered content quality assessment", {})
    
    async def _generate_topic_analysis_section(self, data: Dict[str, Any]) -> ReportSection:
        return ReportSection("Topic Analysis", "Comprehensive topic and theme analysis", {})
    
    async def _generate_content_gaps_section(self, data: Dict[str, Any]) -> ReportSection:
        return ReportSection("Content Gaps", "Identification of content opportunities", {})
    
    async def _generate_social_overview_section(self, data: Dict[str, Any]) -> ReportSection:
        return ReportSection("Social Media Overview", "Social media landscape overview", {})
    
    async def _generate_engagement_analysis_section(self, data: Dict[str, Any]) -> ReportSection:
        return ReportSection("Engagement Analysis", "Social media engagement pattern analysis", {})
    
    async def _generate_viral_content_section(self, data: Dict[str, Any]) -> ReportSection:
        return ReportSection("Viral Content Analysis", "Analysis of high-performing viral content", {})

    async def _generate_custom_section(self, section_name: str, data: Dict[str, Any]) -> ReportSection:
        """Generate custom section."""
        return ReportSection(
            title=section_name,
            content=f"Custom analysis section: {section_name}",
            data={"section_type": "custom"}
        )

    def _format_data_for_table(self, data: Dict[str, Any]) -> List[List[str]]:
        """Format data dictionary for table display."""
        if not data:
            return []
        
        table_data = [["Key", "Value"]]
        for key, value in data.items():
            if isinstance(value, (list, dict)):
                value_str = json.dumps(value)[:100] + "..." if len(json.dumps(value)) > 100 else json.dumps(value)
            else:
                value_str = str(value)
            table_data.append([str(key), value_str])
        
        return table_data

    def _generate_report_id(self) -> str:
        """Generate unique report ID."""
        import uuid
        return str(uuid.uuid4())[:8]

# Global instance
advanced_reporting_service = AdvancedReportingService()